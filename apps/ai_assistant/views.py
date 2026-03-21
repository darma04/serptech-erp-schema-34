"""
==========================================================================
 AI ASSISTANT VIEWS - Chat API & Halaman Pengaturan CRUD
==========================================================================
 Endpoint:
 - POST /ai/chat/         → API chat AI (AJAX)
 - GET  /ai/pengaturan/   → Halaman pengaturan AI Assistant

 Arsitektur Aman:
 User Chat → Backend deteksi intent → ORM ambil data →
 Kirim ringkasan ke AI → AI merapikan teks → Return ke user
==========================================================================
"""
import json
import logging
import ssl
import urllib.request
import urllib.error

# Import dari framework Django
from django.http import JsonResponse
# Import dari framework Django
from django.views import View
# Import dari framework Django
from django.views.generic import TemplateView
# Import dari framework Django
from django.contrib.auth.decorators import login_required
# Import dari framework Django
from django.contrib.auth.mixins import LoginRequiredMixin
# Import dari framework Django
from django.utils.decorators import method_decorator
# Import dari framework Django
from django.views.decorators.csrf import csrf_protect
# Import dari framework Django
from django.contrib import messages
# Import dari framework Django
from django.shortcuts import redirect

# Import dari modul internal proyek
from .models import AIAssistantConfig, ChatHistory, ChatFeedback
# Import dari modul internal proyek
from .intents import detect_intent, gather_data
from web_project import TemplateLayout
from django.db import transaction

logger = logging.getLogger(__name__)


def _get_ssl_context():
    """Buat SSL context yang kompatibel dengan Windows."""
    try:
        ctx = ssl.create_default_context()
        return ctx
    # Tangkap error Exception — lanjutkan tanpa crash
    except Exception:
        # Fallback: skip SSL verification (untuk development)
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        return ctx


# ═══════════════════════════════════════════════════════════════
# SYSTEM PROMPT — Konteks ERP untuk AI
# ═══════════════════════════════════════════════════════════════
SYSTEM_PROMPT = """Kamu adalah AI Business Intelligence Assistant profesional untuk sistem ERP SERPTECH.
Tugasmu: menganalisa data bisnis, membuat laporan, memberikan insight strategis, dan rekomendasi aksi.

ATURAN UTAMA:
1. Gunakan Bahasa Indonesia profesional dan mudah dipahami
2. Berikan analisa mendalam + insight + rekomendasi aksi yang konkret
3. WAJIB gunakan tabel markdown untuk data angka/perbandingan:
   | Kolom 1 | Kolom 2 | Kolom 3 |
   |---------|---------|---------|
   | data    | data    | data    |
4. Selalu tampilkan data dalam tabel jika ada >1 item
5. Jika data menunjukkan tren positif, berikan apresiasi + saran scale up
6. Jika ada masalah, berikan saran perbaikan dengan prioritas (Tinggi/Sedang/Rendah)
7. Jangan mengarang data — hanya analisa data yang diberikan
8. Jawab informatif dan komprehensif (maksimal 500 kata)
9. Gunakan emoji untuk memperjelas kategori dan status
10. Untuk laporan meeting/executive summary: gunakan format narasi profesional + tabel
11. Untuk SWOT: buat 4 kategori (S/W/O/T) masing-masing 2-3 poin dalam tabel
12. Untuk forecasting: berikan prediksi dengan confidence level
13. Untuk analisa risiko: gunakan level 🔴Tinggi/🟡Sedang/🟢Rendah
14. Untuk rencana aksi: buat tabel dengan kolom Aksi, Prioritas, Target, Deadline
15. Ikuti INSTRUKSI khusus yang diberikan bersama data

QUICK ACTIONS — LINK NAVIGASI:
Sertakan link ke halaman ERP yang relevan di akhir respons menggunakan format markdown.
Contoh: → [Lihat Daftar Produk](/produk/list/) atau → [Buka POS Kasir](/pos/)
Gunakan peta URL berikut untuk menentukan link yang tepat:

PETA URL HALAMAN ERP:
- Dashboard utama: /
- Daftar Produk: /produk/list/
- Tambah Produk: /produk/tambah/
- Kategori Produk: /produk/kategori/
- Stok Inventory: /inventory/stok/
- Gudang: /inventory/gudang/
- Transfer Stok: /inventory/transfer/
- Adjustment Stok: /inventory/adjustment/
- Supplier: /pembelian/supplier/
- Purchase Order: /pembelian/purchase-order/
- Customer: /penjualan/customer/
- Sales Order: /penjualan/sales-order/
- POS Kasir: /pos/
- Biaya Operasional: /biaya/
- Laporan: /laporan/
- Karyawan HR: /hr/karyawan/
- AI Dashboard: /ai/dashboard/
- AI Pengaturan: /ai/
- Dashboard Fraud: /fraud/
- Daftar Anomali: /fraud/alerts/
- Rekonsiliasi Kas: /fraud/cash/
- Pengaturan Fraud: /fraud/settings/

SELALU sertakan 1-3 link relevan di akhir setiap jawaban dalam format:
📌 **Quick Actions:**
→ [Label Link](/url-terkait/)

KONTEKS WAKTU:
Jika user bertanya tentang periode waktu tertentu (minggu ini, bulan lalu, kemarin, dll),
data yang diberikan SUDAH difilter sesuai periode tersebut. Analisa sesuai periode yang diminta.

KAPABILITAS:
- Laporan meeting otomatis, Executive summary, Analisa SWOT
- Forecasting/prediksi, Analisa risiko, Rencana aksi
- Perbandingan periode, Stok kritis, Analisa margin produk
- Rekomendasi restock, bundling, strategi harga
- Analisa pelanggan: top customer, customer tidak aktif, frekuensi beli

KONTEKS ERP:
- Modul: Produk, Inventory, Pembelian (PO), Penjualan (SO), POS/Kasir, Biaya, Laporan, HR, Automasi, Fraud Detection
- Mata uang: Rupiah (IDR), Multi-gudang, Multi-metode pembayaran
"""


def _call_gemini(api_key, model, prompt, system_prompt, config):
    """Panggil Google Gemini API.
    Prioritas: google.genai SDK → urllib fallback.
    Mendukung retry otomatis untuk error 429 (rate limit).
    """
    # Coba pakai google.genai SDK (resmi & paling reliable)
    try:
        from google import genai
        return _call_gemini_sdk(genai, api_key, model, prompt, system_prompt, config)
    # Tangkap error ImportError — lanjutkan tanpa crash
    except ImportError:
        logger.info("[AI Gemini] google.genai SDK not installed, using urllib fallback")
        return _call_gemini_urllib(api_key, model, prompt, system_prompt, config)


def _call_gemini_sdk(genai, api_key, model, prompt, system_prompt, config):
    """Panggil Gemini via google.genai SDK resmi."""
    import time

    client = genai.Client(api_key=api_key)
    full_prompt = f"{system_prompt}\\n\\n{prompt}"

    max_retries = 3
    for attempt in range(max_retries):
        # Blok penanganan error — coba jalankan kode di bawah
        try:
            response = client.models.generate_content(
                model=model,
                contents=full_prompt,
                config={
                    "temperature": config.temperature,
                    "max_output_tokens": config.max_tokens,
                }
            )
            if response and response.text:
                return response.text
            logger.warning(f"[AI Gemini SDK] Empty response")
            return None
        # Tangkap error Exception — lanjutkan tanpa crash
        except Exception as e:
            error_str = str(e)
            logger.error(f"[AI Gemini SDK] Error (attempt {attempt+1}): {error_str[:200]}")

            if '429' in error_str and attempt < max_retries - 1:
                wait_time = (attempt + 1) * 5
                logger.info(f"[AI Gemini SDK] Rate limited, waiting {wait_time}s...")
                time.sleep(wait_time)
                continue

            if '429' in error_str:
                raise Exception(
                    "Gemini API: Kuota habis. Solusi:\\n"
                    "1. Buka https://aistudio.google.com/apikeys\\n"
                    "2. Klik project Anda → 'Set up billing'\\n"
                    "3. Tambahkan metode pembayaran (tetap GRATIS, hanya verifikasi)\\n"
                    "4. Setelah billing aktif, kuota gratis menjadi 1500 req/hari"
                )
            raise

    raise Exception("Gemini API: Gagal setelah beberapa percobaan")


def _call_gemini_urllib(api_key, model, prompt, system_prompt, config):
    """Fallback: Panggil Gemini via urllib (tanpa SDK)."""
    import time

    # DIPERBAIKI: API key tidak lagi terekspos di log — disimpan di variabel terpisah
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"

    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": f"{system_prompt}\\n\\n{prompt}"}]
            }
        ],
        "generationConfig": {
            "temperature": config.temperature,
            "maxOutputTokens": config.max_tokens,
        }
    }

    data = json.dumps(payload).encode('utf-8')
    max_retries = 3

    for attempt in range(max_retries):
        req = urllib.request.Request(url, data=data, headers={
            'Content-Type': 'application/json',
        })

        # Blok penanganan error — coba jalankan kode di bawah
        try:
            ssl_ctx = _get_ssl_context()
            with urllib.request.urlopen(req, timeout=60, context=ssl_ctx) as response:
                result = json.loads(response.read().decode('utf-8'))
                candidates = result.get('candidates', [])
                if candidates:
                    parts = candidates[0].get('content', {}).get('parts', [])
                    if parts:
                        return parts[0].get('text', '')
                logger.warning(f"[AI Gemini] No candidates in response")
                return None
        # Tangkap error urllib.error.HTTPError — lanjutkan tanpa crash
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8') if e.fp else ''
            logger.error(f"[AI Gemini] HTTP {e.code} (attempt {attempt+1}): {error_body[:200]}")

            if e.code == 429 and attempt < max_retries - 1:
                wait_time = (attempt + 1) * 5
                time.sleep(wait_time)
                continue

            if e.code == 429:
                raise Exception(
                    "Gemini API: Kuota habis. Solusi:\\n"
                    "1. Buka https://aistudio.google.com/apikeys\\n"
                    "2. Klik project Anda → 'Set up billing'\\n"
                    "3. Tambahkan metode pembayaran (tetap GRATIS, hanya verifikasi)\\n"
                    "4. Setelah billing aktif, kuota gratis menjadi 1500 req/hari"
                )

            # Blok penanganan error — coba jalankan kode di bawah
            try:
                error_data = json.loads(error_body)
                error_msg = error_data.get('error', {}).get('message', '')[:200]
            # Tangkap error Exception — lanjutkan tanpa crash
            except Exception:
                error_msg = error_body[:200]
            raise Exception(f"Gemini API Error {e.code}: {error_msg}")
        # Tangkap error Exception — lanjutkan tanpa crash
        except Exception as e:
            logger.error(f"[AI Gemini] Error: {e}", exc_info=True)
            raise

    raise Exception("Gemini API: Gagal setelah beberapa percobaan (rate limit)")


def _call_openai(api_key, model, prompt, system_prompt, config):
    """Panggil OpenAI ChatGPT API via REST (tanpa library tambahan)."""
    url = "https://api.openai.com/v1/chat/completions"

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": config.max_tokens,
        "temperature": config.temperature,
    }

    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers={
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {api_key}',
    })

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        ssl_ctx = _get_ssl_context()
        with urllib.request.urlopen(req, timeout=45, context=ssl_ctx) as response:
            result = json.loads(response.read().decode('utf-8'))
            choices = result.get('choices', [])
            if choices:
                return choices[0].get('message', {}).get('content', '')
            logger.warning(f"[AI OpenAI] No choices in response: {result}")
            return None
    # Tangkap error urllib.error.HTTPError — lanjutkan tanpa crash
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8') if e.fp else ''
        logger.error(f"[AI OpenAI] HTTP Error {e.code}: {error_body}")
        raise Exception(f"OpenAI API Error {e.code}: {error_body[:200]}")
    # Tangkap error Exception — lanjutkan tanpa crash
    except Exception as e:
        logger.error(f"[AI OpenAI] Error: {e}", exc_info=True)
        raise


def _call_groq(api_key, model, prompt, system_prompt, config):
    """Panggil Groq API (OpenAI-compatible format, GRATIS).
    Endpoint: https://api.groq.com/openai/v1/chat/completions
    Limit: 14.400 req/hari, 30 req/menit (Free Tier)
    """
    url = "https://api.groq.com/openai/v1/chat/completions"

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": config.max_tokens,
        "temperature": config.temperature,
    }

    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers={
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {api_key}',
        'User-Agent': 'SERPTECH-ERP/1.0',
    })

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        ssl_ctx = _get_ssl_context()
        with urllib.request.urlopen(req, timeout=60, context=ssl_ctx) as response:
            result = json.loads(response.read().decode('utf-8'))
            choices = result.get('choices', [])
            if choices:
                return choices[0].get('message', {}).get('content', '')
            logger.warning(f"[AI Groq] No choices in response: {result}")
            return None
    # Tangkap error urllib.error.HTTPError — lanjutkan tanpa crash
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8') if e.fp else ''
        logger.error(f"[AI Groq] HTTP Error {e.code}: {error_body[:300]}")
        # Blok penanganan error — coba jalankan kode di bawah
        try:
            error_data = json.loads(error_body)
            error_msg = error_data.get('error', {}).get('message', error_body[:200])
        # Tangkap error Exception — lanjutkan tanpa crash
        except Exception:
            error_msg = error_body[:200]
        raise Exception(f"Groq API Error {e.code}: {error_msg}")
    # Tangkap error Exception — lanjutkan tanpa crash
    except Exception as e:
        logger.error(f"[AI Groq] Error: {e}", exc_info=True)
        raise


# ═══════════════════════════════════════════════════════════════
# API VIEW — POST /ai/chat/
# ═══════════════════════════════════════════════════════════════

# Wajib login — redirect ke login page jika belum login
@login_required
def ai_chat_api(request):
    """
    API endpoint untuk AI Chat dengan context memory.
    POST body JSON: {"message": "pertanyaan user"}
    Response JSON: {"response": "...", "intent": "...", "source": "...", "chat_id": N}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        body = json.loads(request.body)
        user_message = body.get('message', '').strip()
    # Tangkap error (json.JSONDecodeError, AttributeError) — lanjutkan tanpa crash
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if not user_message:
        return JsonResponse({'error': 'Pesan tidak boleh kosong'}, status=400)

    if len(user_message) > 500:
        return JsonResponse({'error': 'Pesan terlalu panjang (maks 500 karakter)'}, status=400)

    # 1. Muat konfigurasi
    config = AIAssistantConfig.load()

    if not config.aktif:
        return JsonResponse({
            'response': 'AI Assistant sedang nonaktif. Silakan aktifkan di Pengaturan > AI Assistant.',
            'intent': 'system', 'source': 'system'
        })

    # 2. Deteksi intent (multi-intent support)
    intent = detect_intent(user_message)

    # 3. Kumpulkan data via ORM (dengan konteks waktu dari pesan user)
    data = gather_data(intent, user_message)
    ringkasan = data.get('ringkasan', 'Tidak ada data tersedia.')

    # 4. Simpan pesan user ke database
    user_chat = ChatHistory.objects.create(
        user=request.user,
        role='user',
        message=user_message,
        intent=intent,
    )

    # 5. Load context memory — 5 pesan terakhir
    context_messages = ChatHistory.objects.filter(
        user=request.user
    ).order_by('-created_at')[:10]  # 5 pairs = 10 records
    context_messages = list(reversed(context_messages))

    context_text = ""
    if len(context_messages) > 1:  # Ada pesan sebelumnya
        ctx_lines = []
        for msg in context_messages[:-1]:  # Exclude current message
            role_label = "User" if msg.role == 'user' else "AI"
            ctx_lines.append(f"{role_label}: {msg.message[:200]}")
        if ctx_lines:
            context_text = "\\n\\nKONTEKS PERCAKAPAN SEBELUMNYA:\\n" + "\\n".join(ctx_lines[-6:])

    # 6. Panggil AI API
    ai_response = None
    source = 'local'

    if config.api_key:
        # Blok penanganan error — coba jalankan kode di bawah
        try:
            extra_prompt = f"\\n{config.system_prompt}" if config.system_prompt else ""
            full_system = SYSTEM_PROMPT + extra_prompt

            prompt = f"""Pertanyaan user: "{user_message}"
{context_text}

Berikut data dari sistem ERP:
{ringkasan}

Berdasarkan data di atas, berikan analisa atau jawaban yang profesional dan informatif dalam Bahasa Indonesia."""

            if config.provider == 'gemini':
                ai_response = _call_gemini(config.api_key, config.model_name, prompt, full_system, config)
            elif config.provider == 'openai':
                ai_response = _call_openai(config.api_key, config.model_name, prompt, full_system, config)
            elif config.provider == 'groq':
                ai_response = _call_groq(config.api_key, config.model_name, prompt, full_system, config)

            source = config.get_provider_display() if ai_response else 'fallback'

        # Tangkap error Exception — lanjutkan tanpa crash
        except Exception as e:
            logger.error(f"[AI Chat] API Error: {e}", exc_info=True)
            ai_response = f"⚠️ AI Error ({str(e)[:100]})\\n\\nBerikut data mentah:\\n\\n{ringkasan}"
            source = 'fallback'

    if not ai_response:
        ai_response = f"📊 {ringkasan}\\n\\n💡 *Untuk analisa AI yang lebih baik, masukkan API Key di Pengaturan > AI Assistant.*"

    # 7. Simpan respons AI ke database
    ai_chat = ChatHistory.objects.create(
        user=request.user,
        role='assistant',
        message=ai_response,
        intent=intent,
        source=source,
    )

    return JsonResponse({
        'response': ai_response,
        'intent': intent,
        'source': source,
        'chat_id': ai_chat.id,
    })


# ═══════════════════════════════════════════════════════════════
# AUTO INSIGHT — GET /ai/insight/
# ═══════════════════════════════════════════════════════════════

# Wajib login — redirect ke login page jika belum login
@login_required
def auto_insight(request):
    """
    Auto Insight: Ringkasan bisnis otomatis saat user buka chat.
    Tidak memanggil AI API — hanya query ORM langsung.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        # Import dari framework Django
        from django.utils import timezone
        # Import dari framework Django
        from django.db.models import Sum
        today = timezone.now().date()
        month_start = today.replace(day=1)

        # Import dari modul internal proyek
        from apps.penjualan.models import SalesOrder
        # Import dari modul internal proyek
        from apps.pos.models import POSTransaction
        # Import dari modul internal proyek
        from apps.produk.models import Produk, Stok

        # Omzet hari ini
        pos_today = float(POSTransaction.objects.filter(
            status='paid', tanggal__date=today
        ).aggregate(t=Sum('total_harga'))['t'] or 0)
        # Query database — ambil data pos_today_count yang sesuai filter
        pos_today_count = POSTransaction.objects.filter(
            status='paid', tanggal__date=today
        ).count()

        # Omzet bulan ini
        so_month = float(SalesOrder.objects.filter(
            status__in=['confirmed', 'delivered', 'completed'],
            tanggal__date__gte=month_start
        ).aggregate(t=Sum('total_harga'))['t'] or 0)
        # Query database — ambil data pos_month yang sesuai filter
        pos_month = float(POSTransaction.objects.filter(
            status='paid', tanggal__date__gte=month_start
        ).aggregate(t=Sum('total_harga'))['t'] or 0)
        total_month = so_month + pos_month

        # Stok kritis
        ids_with_stock = Stok.objects.filter(jumlah__gt=0).values_list('produk_id', flat=True)
        # Query database — ambil data stok_habis yang sesuai filter
        stok_habis = Produk.objects.filter(aktif=True).exclude(id__in=ids_with_stock).count()

        # Stok rendah
        stok_rendah = Stok.objects.values('produk').annotate(
            total=Sum('jumlah')
        ).filter(total__gt=0, total__lt=10).count()

        insights = []
        insights.append(f"💰 Omzet hari ini: Rp {pos_today:,.0f} ({pos_today_count} trx)")
        insights.append(f"📊 Omzet bulan ini: Rp {total_month:,.0f}")

        if stok_habis > 0:
            insights.append(f"🔴 {stok_habis} produk stok habis!")
        if stok_rendah > 0:
            insights.append(f"🟡 {stok_rendah} produk stok rendah (<10)")
        if stok_habis == 0 and stok_rendah == 0:
            insights.append("🟢 Semua stok tercukupi")

        return JsonResponse({
            'insights': insights,
            'date': today.strftime('%d %B %Y'),
        })

    # Tangkap error Exception — lanjutkan tanpa crash
    except Exception as e:
        logger.error(f"[AI Insight] Error: {e}", exc_info=True)
        return JsonResponse({'insights': [], 'date': ''})


# ═══════════════════════════════════════════════════════════════
# FEEDBACK — POST /ai/feedback/
# ═══════════════════════════════════════════════════════════════

# Wajib login — redirect ke login page jika belum login
@login_required
def chat_feedback(request):
    """Simpan feedback 👍/👎 dari user."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        body = json.loads(request.body)
        feedback_type = body.get('feedback', '').strip()
        chat_id = body.get('chat_id')
        message_text = body.get('message_text', '')[:500]
    # Tangkap error (json.JSONDecodeError, AttributeError) — lanjutkan tanpa crash
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if feedback_type not in ('up', 'down'):
        return JsonResponse({'error': 'Feedback harus "up" atau "down"'}, status=400)

    chat_obj = None
    if chat_id:
        # Blok penanganan error — coba jalankan kode di bawah
        try:
            # Query database — ambil satu data chat_obj
            chat_obj = ChatHistory.objects.get(id=chat_id, user=request.user)
        # Tangkap error ChatHistory.DoesNotExist — lanjutkan tanpa crash
        except ChatHistory.DoesNotExist:
            pass

    ChatFeedback.objects.create(
        user=request.user,
        chat=chat_obj,
        feedback=feedback_type,
        message_text=message_text,
    )

    return JsonResponse({'status': 'ok', 'message': 'Terima kasih atas feedback Anda!'})


# ═══════════════════════════════════════════════════════════════
# HISTORY — GET /ai/history/
# ═══════════════════════════════════════════════════════════════

# Wajib login — redirect ke login page jika belum login
@login_required
def chat_history_api(request):
    """Load riwayat chat user dari database."""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Query database — ambil data messages_qs yang sesuai filter
    messages_qs = ChatHistory.objects.filter(
        user=request.user
    ).order_by('-created_at')[:50]  # Last 50 messages

    messages_list = []
    for msg in reversed(messages_qs):
        messages_list.append({
            'id': msg.id,
            'role': msg.role,
            'message': msg.message,
            'intent': msg.intent,
            'source': msg.source,
            'time': msg.created_at.strftime('%H:%M'),
        })

    return JsonResponse({'messages': messages_list})


# ═══════════════════════════════════════════════════════════════
# CLEAR HISTORY — POST /ai/clear/
# ═══════════════════════════════════════════════════════════════

# Wajib login — redirect ke login page jika belum login
@login_required
def clear_history(request):
    """Hapus riwayat chat user."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Query database — ambil data deleted_count yang sesuai filter
    deleted_count = ChatHistory.objects.filter(user=request.user).delete()[0]
    return JsonResponse({
        'status': 'ok',
        'message': f'{deleted_count} pesan dihapus.',
    })


# ═══════════════════════════════════════════════════════════════
# AI DASHBOARD — Halaman Analytics /ai/dashboard/
# ═══════════════════════════════════════════════════════════════

class AIDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard AI dengan skor kesehatan bisnis, prediksi, dan anomali."""
    template_name = 'ai_assistant/dashboard.html'

    # Menambahkan data konteks tambahan ke template
    def get_context_data(self, **kwargs):
        context = TemplateLayout.init(self, super().get_context_data(**kwargs))

        import calendar
        # Import dari framework Django
        from django.utils import timezone
        # Import dari framework Django
        from django.db.models import Sum, Avg, F
        from datetime import timedelta, date

        # Default values agar template tidak error jika data gagal dimuat
        defaults = {
            'rev_now': 0, 'rev_prev': 0, 'rev_growth': 0,
            'total_trx': 0, 'biaya_now': 0, 'profit': 0,
            'total_produk': 0, 'stok_habis': 0, 'stok_rendah': 0, 'stok_sehat': 0,
            'margin_neg': 0, 'avg_margin': 0, 'po_pending': 0, 'biaya_ratio': 0,
            'health_score': 50, 'health_level': 'Memuat...', 'health_color': '#999',
            'anomalies': [], 'monthly_labels': '[]', 'monthly_values': '[]',
            'total_chats': 0, 'total_feedback_up': 0, 'total_feedback_down': 0,
        }
        context.update(defaults)

        today = timezone.now().date()
        month_start = today.replace(day=1)
        prev_end = month_start - timedelta(days=1)
        prev_start = prev_end.replace(day=1)

        # Blok penanganan error — coba jalankan kode di bawah
        try:
            # Import dari modul internal proyek
            from apps.penjualan.models import SalesOrder
            # Import dari modul internal proyek
            from apps.pos.models import POSTransaction
            # Import dari modul internal proyek
            from apps.produk.models import Produk, Stok
            # Import dari modul internal proyek
            from apps.biaya.models import TransaksiBiaya
            # Import dari modul internal proyek
            from apps.pembelian.models import PurchaseOrder

            # ── Helper: hitung revenue antara 2 tanggal ──
            def get_revenue(start, end):
                # Query database — ambil data so yang sesuai filter
                so = float(SalesOrder.objects.filter(
                    status__in=['confirmed', 'delivered', 'completed'],
                    tanggal__date__gte=start, tanggal__date__lte=end
                ).aggregate(t=Sum('total_harga'))['t'] or 0)
                # Query database — ambil data pos yang sesuai filter
                pos = float(POSTransaction.objects.filter(
                    status='paid', tanggal__date__gte=start, tanggal__date__lte=end
                ).aggregate(t=Sum('total_harga'))['t'] or 0)
                return so + pos

            # ── REVENUE ──
            rev_now = get_revenue(month_start, today)
            rev_prev = get_revenue(prev_start, prev_end)
            rev_growth = round((rev_now - rev_prev) / rev_prev * 100, 1) if rev_prev > 0 else 0

            # ── TRANSAKSI ──
            trx_so = SalesOrder.objects.filter(
                status__in=['confirmed', 'delivered', 'completed'],
                tanggal__date__gte=month_start
            ).count()
            # Query database — ambil data trx_pos yang sesuai filter
            trx_pos = POSTransaction.objects.filter(
                status='paid', tanggal__date__gte=month_start
            ).count()
            total_trx = trx_so + trx_pos

            # ── BIAYA ──
            biaya_now = float(TransaksiBiaya.objects.filter(
                tanggal__gte=month_start
            ).aggregate(t=Sum('jumlah'))['t'] or 0)
            profit = rev_now - biaya_now

            # ── STOK (evaluate list() agar tidak lazy subquery) ──
            total_produk = Produk.objects.filter(aktif=True).count()
            ids_with_stock = list(
                # Query database — ambil data Stok.objects.filter(jumlah__gt yang sesuai filter
                Stok.objects.filter(jumlah__gt=0)
                .values_list('produk_id', flat=True)
                .distinct()
            )
            # Query database — ambil data stok_habis yang sesuai filter
            stok_habis = Produk.objects.filter(aktif=True).exclude(
                id__in=ids_with_stock
            ).count()
            stok_rendah = Stok.objects.values('produk').annotate(
                total=Sum('jumlah')
            ).filter(total__gt=0, total__lt=10).count()
            stok_sehat = max(0, total_produk - stok_habis - stok_rendah)

            # ── MARGIN ──
            margin_neg = Produk.objects.filter(
                aktif=True, harga_jual__lt=F('harga_beli')
            ).count()
            # Query database — ambil data avg_margin_val yang sesuai filter
            avg_margin_val = Produk.objects.filter(
                aktif=True, harga_jual__gt=0
            ).aggregate(
                avg_margin=Avg(
                    (F('harga_jual') - F('harga_beli')) * 100 / F('harga_jual')
                )
            )['avg_margin']
            avg_margin = round(float(avg_margin_val or 0), 1)

            # ── PO ──
            po_pending = PurchaseOrder.objects.filter(status='draft').count()

            # ═══ SKOR KESEHATAN BISNIS (0-100) ═══
            score = 50
            if rev_growth > 10: score += 15
            elif rev_growth > 0: score += 10
            elif rev_growth > -10: score -= 5
            else: score -= 15

            if avg_margin > 20: score += 15
            elif avg_margin > 10: score += 10
            elif avg_margin > 0: score += 5
            else: score -= 10

            stok_ratio = stok_habis / max(total_produk, 1) * 100
            if stok_ratio == 0: score += 10
            elif stok_ratio < 20: score += 5
            elif stok_ratio < 50: score -= 5
            else: score -= 10

            biaya_ratio = biaya_now / max(rev_now, 1) * 100
            if biaya_ratio < 30: score += 10
            elif biaya_ratio < 50: score += 5
            elif biaya_ratio < 70: score -= 5
            else: score -= 10

            score = max(0, min(100, score))

            if score >= 80: health_level, health_color = 'Sangat Sehat', '#28a745'
            elif score >= 60: health_level, health_color = 'Sehat', '#4caf50'
            elif score >= 40: health_level, health_color = 'Perlu Perhatian', '#ff9800'
            elif score >= 20: health_level, health_color = 'Bermasalah', '#f44336'
            else: health_level, health_color = 'Kritis', '#d32f2f'

            # ═══ DETEKSI ANOMALI ═══
            anomalies = []
            if stok_habis > 0:
                anomalies.append({
                    'level': 'danger', 'icon': 'ri-error-warning-line',
                    'title': f'{stok_habis} Produk Stok Habis',
                    'desc': 'Segera lakukan restock untuk menghindari kehilangan penjualan.',
                })
            if margin_neg > 0:
                anomalies.append({
                    'level': 'warning', 'icon': 'ri-funds-line',
                    'title': f'{margin_neg} Produk Margin Negatif',
                    'desc': 'Harga jual lebih rendah dari harga beli. Perlu evaluasi harga.',
                })
            if rev_growth < -20:
                anomalies.append({
                    'level': 'danger', 'icon': 'ri-arrow-down-line',
                    'title': f'Revenue Turun {abs(rev_growth)}%',
                    'desc': 'Penjualan bulan ini turun signifikan dibanding bulan lalu.',
                })
            if biaya_ratio > 60:
                anomalies.append({
                    'level': 'warning', 'icon': 'ri-money-dollar-circle-line',
                    'title': f'Rasio Biaya Tinggi ({biaya_ratio:.0f}%)',
                    'desc': 'Biaya operasional melebihi 60% dari revenue.',
                })
            if stok_rendah > 3:
                anomalies.append({
                    'level': 'info', 'icon': 'ri-inbox-line',
                    'title': f'{stok_rendah} Produk Stok Rendah',
                    'desc': 'Pertimbangkan untuk restock sebelum habis.',
                })
            if po_pending > 5:
                anomalies.append({
                    'level': 'info', 'icon': 'ri-file-list-line',
                    'title': f'{po_pending} PO Menunggu Approval',
                    'desc': 'Ada purchase order yang belum di-approve.',
                })
            if not anomalies:
                anomalies.append({
                    'level': 'success', 'icon': 'ri-check-double-line',
                    'title': 'Tidak Ada Anomali',
                    'desc': 'Semua metrik bisnis dalam kondisi normal.',
                })

            # ═══ TREN BULANAN (6 bulan — pakai calendar untuk akurasi) ═══
            monthly_labels = []
            monthly_values = []
            cur_year, cur_month = today.year, today.month
            for i in range(5, -1, -1):
                m = cur_month - i
                y = cur_year
                while m <= 0:
                    m += 12
                    y -= 1
                m_start = date(y, m, 1)
                _, last_day = calendar.monthrange(y, m)
                m_end = date(y, m, last_day) if i > 0 else today
                m_rev = get_revenue(m_start, m_end)
                monthly_labels.append(m_start.strftime('%b %Y'))
                monthly_values.append(m_rev)

            # ═══ AI USAGE STATS ═══
            total_chats = ChatHistory.objects.filter(user=self.request.user).count()
            # Query database — ambil data total_feedback_up yang sesuai filter
            total_feedback_up = ChatFeedback.objects.filter(
                user=self.request.user, feedback='up'
            ).count()
            # Query database — ambil data total_feedback_down yang sesuai filter
            total_feedback_down = ChatFeedback.objects.filter(
                user=self.request.user, feedback='down'
            ).count()

            context.update({
                'rev_now': rev_now,
                'rev_prev': rev_prev,
                'rev_growth': rev_growth,
                'total_trx': total_trx,
                'biaya_now': biaya_now,
                'profit': profit,
                'total_produk': total_produk,
                'stok_habis': stok_habis,
                'stok_rendah': stok_rendah,
                'stok_sehat': stok_sehat,
                'margin_neg': margin_neg,
                'avg_margin': avg_margin,
                'po_pending': po_pending,
                'biaya_ratio': round(biaya_ratio, 1),
                'health_score': score,
                'health_level': health_level,
                'health_color': health_color,
                'anomalies': anomalies,
                'monthly_labels': json.dumps(monthly_labels),
                'monthly_values': json.dumps(monthly_values),
                'total_chats': total_chats,
                'total_feedback_up': total_feedback_up,
                'total_feedback_down': total_feedback_down,
            })

        # Tangkap error Exception — lanjutkan tanpa crash
        except Exception as e:
            logger.error(f"[AI Dashboard] Error: {e}", exc_info=True)
            # Data konteks: dashboard_error — untuk ditampilkan di template
            context['dashboard_error'] = str(e)

        return context


# ═══════════════════════════════════════════════════════════════
# PENGATURAN VIEW — CRUD Halaman Pengaturan AI Assistant
# ═══════════════════════════════════════════════════════════════

class AIAssistantSettingsView(LoginRequiredMixin, TemplateView):
    """Halaman pengaturan AI Assistant (GET = tampilkan, POST = simpan)."""
    template_name = 'ai_assistant/index.html'

    # Menambahkan data konteks tambahan ke template
    def get_context_data(self, **kwargs):
        # TemplateLayout.init → menyuntikkan is_menu, is_navbar, menu_data, dll
        context = TemplateLayout.init(self, super().get_context_data(**kwargs))
        # Data konteks: config — untuk ditampilkan di template
        context['config'] = AIAssistantConfig.load()
        # Data konteks: provider_choices — untuk ditampilkan di template
        context['provider_choices'] = AIAssistantConfig.PROVIDER_CHOICES
        # Data konteks: page_title — untuk ditampilkan di template
        context['page_title'] = 'AI Assistant'
        return context

    # Handle HTTP POST request
    def post(self, request, *args, **kwargs):
        config = AIAssistantConfig.load()

        config.provider = request.POST.get('provider', 'gemini')
        config.api_key = request.POST.get('api_key', '').strip()
        config.model_name = request.POST.get('model_name', 'gemini-2.0-flash').strip()
        config.aktif = request.POST.get('aktif') == 'on'
        config.ai_name = request.POST.get('ai_name', 'AI Assistant').strip() or 'AI Assistant'
        config.system_prompt = request.POST.get('system_prompt', '').strip()

        # Blok penanganan error — coba jalankan kode di bawah
        try:
            config.max_tokens = int(request.POST.get('max_tokens', 1024))
        # Tangkap error (ValueError, TypeError) — lanjutkan tanpa crash
        except (ValueError, TypeError):
            config.max_tokens = 1024

        # Blok penanganan error — coba jalankan kode di bawah
        try:
            config.temperature = float(request.POST.get('temperature', 0.7))
        # Tangkap error (ValueError, TypeError) — lanjutkan tanpa crash
        except (ValueError, TypeError):
            config.temperature = 0.7

        config.save()
        # Tampilkan pesan sukses ke user
        messages.success(request, 'Pengaturan AI Assistant berhasil disimpan!')
        # Redirect ke halaman tujuan
        return redirect('ai_assistant:index')


    # ═══════════════════════════════════════════════════════════════
    # EXPORT ENDPOINTS — PDF & Excel
    # ═══════════════════════════════════════════════════════════════

    # Wajib login — redirect ke login page jika belum login
@login_required
def export_chat_pdf(request):
    """Export riwayat chat ke file PDF."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        body = json.loads(request.body)
        chat_messages = body.get('messages', [])
    # Tangkap error (json.JSONDecodeError, AttributeError) — lanjutkan tanpa crash
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if not chat_messages:
        return JsonResponse({'error': 'Tidak ada pesan untuk di-export'}, status=400)

    # Import dari framework Django
    from django.http import HttpResponse
    from datetime import datetime
    import io

    # Generate PDF menggunakan format sederhana (tanpa library tambahan)
    # Menggunakan format teks + header yang bisa dibuka di PDF reader
    buffer = io.BytesIO()

    # Buat PDF sederhana secara manual (PDF 1.4 compliant)
    pdf_lines = []
    pdf_lines.append('%PDF-1.4')

    # Konversi teks chat ke content
    content_lines = []
    content_lines.append('RIWAYAT CHAT AI ASSISTANT - SERPTECH ERP')
    content_lines.append('=' * 50)
    content_lines.append(f'Tanggal Export: {datetime.now().strftime("%d %B %Y, %H:%M WIB")}')
    content_lines.append(f'User: {request.user.get_full_name() or request.user.username}')
    content_lines.append(f'Total Pesan: {len(chat_messages)}')
    content_lines.append('')
    content_lines.append('-' * 50)

    for msg in chat_messages:
        role = 'ANDA' if msg.get('type') == 'user' else 'AI ASSISTANT'
        waktu = msg.get('time', '')
        teks = msg.get('text', '')
        content_lines.append(f'')
        content_lines.append(f'[{waktu}] {role}:')
        # Wrap long lines
        for line in teks.split('\\n'):
            while len(line) > 80:
                content_lines.append(f'  {line[:80]}')
                line = line[80:]
            content_lines.append(f'  {line}')
        content_lines.append('')

    content_lines.append('-' * 50)
    content_lines.append('Dihasilkan oleh SERPTECH ERP AI Assistant')

    # Encode content ke PDF text stream
    text_content = '\\n'.join(content_lines)

    # Gunakan format yang lebih kompatibel: Plain text PDF
    stream_content = f"""BT
    /F1 10 Tf
    40 800 Td
    12 TL
    """
    for line in content_lines:
        # Escape karakter khusus PDF
        safe_line = line.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')
        stream_content += f"({safe_line}) Tj T*\\n"

    stream_content += "ET"
    stream_bytes = stream_content.encode('latin-1', errors='replace')

    # Build PDF objects
    objects = []

    # Object 1: Catalog
    objects.append('1 0 obj\\n<< /Type /Catalog /Pages 2 0 R >>\\nendobj')

    # Object 2: Pages
    objects.append('2 0 obj\\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\\nendobj')

    # Object 3: Page
    page_height = max(842, len(content_lines) * 14 + 100)
    objects.append(f'3 0 obj\\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 {page_height}] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\\nendobj')

    # Object 4: Content Stream
    # Reset posisi Y berdasarkan tinggi halaman
    stream_content_final = f"""BT
    /F1 10 Tf
    40 {page_height - 40} Td
    12 TL
    """
    for line in content_lines:
        safe_line = line.replace('\\\\', '\\\\\\\\').replace('(', '\\\\(').replace(')', '\\\\)')
        stream_content_final += f"({safe_line}) Tj T*\\n"
    stream_content_final += "ET"
    stream_bytes_final = stream_content_final.encode('latin-1', errors='replace')

    objects.append(f'4 0 obj\\n<< /Length {len(stream_bytes_final)} >>\\nstream\\n'.encode('latin-1') + stream_bytes_final + b'\\nendstream\\nendobj')

    # Object 5: Font
    objects.append('5 0 obj\\n<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>\\nendobj')

    # Build PDF
    pdf_output = b'%PDF-1.4\\n'
    offsets = []
    for i, obj in enumerate(objects):
        offsets.append(len(pdf_output))
        if isinstance(obj, bytes):
            pdf_output += obj + b'\\n'
        else:
            pdf_output += obj.encode('latin-1') + b'\\n'

    # Cross-reference table
    xref_offset = len(pdf_output)
    pdf_output += b'xref\\n'
    pdf_output += f'0 {len(objects) + 1}\\n'.encode()
    pdf_output += b'0000000000 65535 f \\n'
    for offset in offsets:
        pdf_output += f'{offset:010d} 00000 n \\n'.encode()

    pdf_output += b'trailer\\n'
    pdf_output += f'<< /Size {len(objects) + 1} /Root 1 0 R >>\\n'.encode()
    pdf_output += b'startxref\\n'
    pdf_output += f'{xref_offset}\\n'.encode()
    pdf_output += b'%%EOF'

    response = HttpResponse(pdf_output, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="AI_Chat_SERPTECH_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return response


# Wajib login -- redirect ke login page jika belum login
@login_required
def export_chat_excel(request):
    """Export riwayat chat ke file Excel (.xlsx)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        body = json.loads(request.body)
        chat_messages = body.get('messages', [])
    # Tangkap error (json.JSONDecodeError, AttributeError) — lanjutkan tanpa crash
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    if not chat_messages:
        return JsonResponse({'error': 'Tidak ada pesan untuk di-export'}, status=400)

    # Import dari framework Django
    from django.http import HttpResponse
    from datetime import datetime
    import io

    # Blok penanganan error — coba jalankan kode di bawah
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    # Tangkap error ImportError — lanjutkan tanpa crash
    except ImportError:
        # Fallback ke CSV jika openpyxl tidak tersedia
        return _export_chat_csv(request, chat_messages)

    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Chat AI"

    # Styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="666CFF", end_color="666CFF", fill_type="solid")
    sub_font = Font(size=10, italic=True, color="666666")
    col_header_font = Font(bold=True, size=11, color="FFFFFF")
    col_header_fill = PatternFill(start_color="7C5CBF", end_color="7C5CBF", fill_type="solid")
    user_fill = PatternFill(start_color="E8E8FF", end_color="E8E8FF", fill_type="solid")
    bot_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')

    # Header
    ws.merge_cells('A1:C1')
    ws['A1'] = 'Riwayat Chat AI Assistant — SERPTECH ERP'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:C2')
    ws['A2'] = f'Export: {datetime.now().strftime("%d %B %Y, %H:%M WIB")} | User: {request.user.get_full_name() or request.user.username}'
    ws['A2'].font = sub_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Column Headers
    headers = ['Waktu', 'Pengirim', 'Pesan']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 80

    # Data rows
    for i, msg in enumerate(chat_messages):
        row = i + 5
        role = 'Anda' if msg.get('type') == 'user' else 'AI Assistant'
        fill = user_fill if msg.get('type') == 'user' else bot_fill

        cell_time = ws.cell(row=row, column=1, value=msg.get('time', ''))
        cell_role = ws.cell(row=row, column=2, value=role)
        cell_text = ws.cell(row=row, column=3, value=msg.get('text', ''))

        for cell in [cell_time, cell_role, cell_text]:
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = wrap_align

        cell_role.font = Font(bold=True, color="666CFF" if msg.get('type') == 'user' else "7C5CBF")

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="AI_Chat_SERPTECH_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    return response


def _export_chat_csv(request, chat_messages):
    """Fallback export ke CSV jika openpyxl tidak tersedia."""
    from django.http import HttpResponse
    from datetime import datetime
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Waktu', 'Pengirim', 'Pesan'])

    for msg in chat_messages:
        role = 'Anda' if msg.get('type') == 'user' else 'AI Assistant'
        writer.writerow([msg.get('time', ''), role, msg.get('text', '')])

    response = HttpResponse(buffer.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="AI_Chat_SERPTECH_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
    return response
